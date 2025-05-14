import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook


def extract_column_from_sheet(file_path: str, sheet_name: str, input_columns: list, additional_columns: list = None):
    """
    Extracts a column from a specific sheet in an Excel file.

    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str): Sheet name to read.
        column_name (str): Column name to extract.

    Returns:
        dict: Dictionary of column name to list of values.
    
    Raises:
        ValueError: If the sheet or column is not found.
        FileNotFoundError: If the file does not exist.
    """
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")

    missing_columns = [col for col in input_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Columns not found in sheet '{sheet_name}': {missing_columns}")
        
    extracted_data = {
        col: df[col].tolist()
        for col in input_columns
    }

    if additional_columns is not None:
        process_additional_columns(extracted_data, additional_columns)

    return extracted_data

def process_additional_columns(data, additional_columns):
    for column in additional_columns:
        #Process Borehole additional columns
        if column == "* Date [-]":
            data[column] = "0"
        if column == "* Time [-]":
            data[column] = "12:00:00 PM"
        if column == "Method [-]":
            data[column] = "0"
        if column =="Equipment [-]":
            data[column] = "0"
        if column =="Core diameter [mm]":
            data[column] = "0"
        if column =="Company [-]":
            data[column] = "0"

        #Process SoilTest additional columns
        #["Test Date [DD-MM-YYYY]","AGS Code","Unit","Accuracy"]
        if column =="Test Date [DD-MM-YYYY]":
            data[column] = "0"
        if column =="AGS Code":
            data[column] = "=INDEX(Test_results[[#All],[AGS Code]],MATCH(Pointdata[@[Parameter]],Test_results[[#All],[Test name]],0))"
        if column =="Unit":
            data[column] = "=INDEX(Test_results[[#All],[Unit]],MATCH(Pointdata[@[Parameter]],Test_results[[#All],[Test name]],0))"
        if column =="Accuracy":
            data[column] = "=INDEX(Test_results[[#All],[Accuracy]],MATCH(Pointdata[@[Parameter]],Test_results[[#All],[Test name]],0))"

def process_excel(input_file_path: str, column_map: dict, reference: dict):
    """
    Write data to a new Excel file.
    lithology
    ~ Bore column (kiri) => Name column (kanan)
    ~ Depth1 column (kiri) => Depth top column (kanan)
    ~ Depth2 column (kiri) => Depth bottom column (kanan)
    ~ Keyword column (kiri) => Lithology column (kanan)
    ~ comment column (kiri) => Remarks lithology (kanan)

    Args:
    - output_file (str): Path of the output Excel file.
    - data (list): Data to write to the Excel sheet.
    - sheet_name (str): Name of the sheet to write to.
    - column_name (str): Name of the column in the sheet.
    """
    data_dictionary = extract_column_from_sheet(file_path=input_file_path, sheet_name=reference.get("input_sheet"), input_columns=column_map.keys(), additional_columns=reference.get("additional_columns", None))
    df = pd.DataFrame(data_dictionary)
    df.rename(columns=column_map, inplace=True)

    write_excel_table(df, output_path=reference.get('output_path'), sheet_name=reference.get('output_sheet'))
        
    if reference.get("additional_table") is not None :
        additional_data = extract_column_from_sheet(file_path=reference.get("additional_input_file"), sheet_name=reference.get("additional_input_sheet"), input_columns=reference.get("additional_input_columns"))
        adf = pd.DataFrame(additional_data)
        wb = load_workbook(reference.get('output_path'))
        ws = wb.create_sheet(title="List")
        for r in dataframe_to_rows(adf, index=False, header=True):
            ws.append(r)
        max_row = ws.max_row
        max_col = ws.max_column
        end_col = ws.cell(row=1, column=max_col).column_letter
        table_range = f"A1:{end_col}{max_row}"
        table = Table(displayName="Test_results", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        wb.save(reference.get('output_path'))

def write_excel_table(df, output_path="output1.xlsx", sheet_name="Sheet1", table_name="Table1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    max_row = ws.max_row
    max_col = ws.max_column
    end_col = ws.cell(row=1, column=max_col).column_letter
    table_range = f"A1:{end_col}{max_row}"
    table = Table(displayName=table_name, ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    wb.save(output_path)

def additional_sheet():
   pass

def lithology_extract(input_file):
    input_columns = ["Bore", "Depth1", "Depth2", "Keyword", "Comment"]
    output_columns = ["* Name [-]", "Depth top [m]", "Depth bottom [m]", "Lithology [-]", "Remarks lithology [-]"]
    column_map = dict(zip(input_columns, output_columns))

    reference = {
        "input_sheet":"Lithology",
        "output_sheet":"INPUT Lithology",
        "output_path":"Lithology_Input_Template_rev2.xlsx"
    }
    process_excel(input_file, column_map, reference)

def borehole_extract(input_file):
    input_columns = ["Bore","Enabled","Easting","Northing","Elevation","TotalDepth","CollarElevation","Comments"]
    output_columns = ["* Name [-]","* Easting [m]","* Northing [m]","* Elevation [m reference]","* Ground water table [m below elevation]","Notes [-]"]
    column_map = dict(zip(input_columns, output_columns))

    additional_columns = ["* Date [-]","* Time [-]","Method [-]","Equipment [-]","Core diameter [mm]","Company [-]"]

    reference = {
        "additional_columns":additional_columns,
        "input_sheet":"Location",
        "output_sheet":"INPUT Borehole",
        "output_path":"Borehole_Base_Input_Template_rev2.xlsx"
    }
    process_excel(input_file, column_map, reference)

def soil_extract(input_file):
    input_columns = ["Bore", "Depth1", "Depth2", "Name", "Value"]
    output_columns = ["Investigation Point", "Depth top [m]", "Depth Bottom [m]", "Parameter", "Test Result"]
    column_map = dict(zip(input_columns, output_columns))

    additional_columns = ["Test Date [DD-MM-YYYY]","AGS Code","Unit","Accuracy"]

    reference = {
        "additional_columns":additional_columns,
        "input_sheet":"Interval",
        "output_sheet":"INPUT Soil Test",
        "output_path":"Soil_Test_Input_Template_rev2.xlsx",
        "additional_input_file":"src\\input\\Soil Test List.xlsx",
        "additional_input_sheet":"List",
        "additional_table":"Test_results",
        "additional_input_columns": ["Test name", "AGS Code", "Unit", "Accuracy", "Type (for sorting)", "Remarks"]
    }
    process_excel(input_file, column_map, reference)

def test():
    input_file = "C:\\Workspace\\gtatool\\src\\input\\203935 Grimsby Riverside - Soil data 1.xlsx"
    lithology_extract(input_file)
    borehole_extract(input_file)
    soil_extract(input_file)

if __name__ == "__main__":
    test()